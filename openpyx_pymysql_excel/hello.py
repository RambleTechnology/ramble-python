# 导入load_workbook函数，用于加载Excel文件
from openpyxl import load_workbook
# 导入pymysql库，用于连接和操作MySQL数据库  
import pymysql  

# 连接到MySQL数据库
mydb = pymysql.connect(
    host="192.168.1.32",
    user="root",
    passwd="******",
    db="cml-temp"
)

# 打开Excel文件
wb = load_workbook(filename=r'./data.xlsx')  # 加载Excel文件
sheet = wb.active  # 获取活动工作表

# 获取表头
header = [cell.value for cell in sheet[1]]  # 获取第一行的单元格值，作为表头

# 遍历每一行数据，并将其插入到数据库中
cursor = mydb.cursor()  # 创建游标对象，用于执行SQL语句
count = 0  # 计数器，用于记录插入的数据条数
for row in sheet.iter_rows(min_row=2, values_only=True):  # 从第二行开始遍历每一行数据
    # way1:读取excel的表头作为数据库的列
    # sql = f"INSERT INTO car ({', '.join(header)}) VALUES ({', '.join(['%s'] * len(header))})"

    # way2:自定义表列
    sql = f"INSERT INTO car (id,name,brand) VALUES ({', '.join(['%s'] * len(header))})"

    # 构建插入数据的SQL语句，使用占位符%s
    cursor.execute(sql, row)  # 执行SQL语句，将行数据插入到数据库中
    count += 1  # 每插入一条数据，计数器加1
    print(f"正在插入{count}条数据")  # 输出插入的数据条数

# 提交更改并关闭数据库连接
mydb.commit()  # 提交对数据库的更改
cursor.close()  # 关闭游标对象
mydb.close()  # 关闭与数据库的连接